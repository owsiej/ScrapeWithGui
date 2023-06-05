import pandas as pd
import io
from UliPlot.XLSX import auto_adjust_xlsx_column_width
from datetime import datetime
from openpyxl.styles import Alignment


def create_memory_excel_file(listOfFlatsByDeveloper: list):
    devs_names = [dev['developerName']
                  for dev in listOfFlatsByDeveloper]
    flatsDataframes = [pd.DataFrame(flats["flatsData"])
                       for flats in listOfFlatsByDeveloper]
    summaryDataframe = create_summary_to_excel(flatsDataframes, devs_names)
    memory_file = io.BytesIO()
    with pd.ExcelWriter(memory_file, engine="openpyxl") as writer:
        summaryDataframe.to_excel(writer, sheet_name="Summary", startrow=1)
        auto_adjust_xlsx_column_width(summaryDataframe, writer, sheet_name="Summary", margin=1)
        writer.sheets['Summary'].merge_cells(start_row=1, start_column=1, end_row=1,
                                             end_column=10)
        writer.sheets['Summary'].cell(row=1, column=1).value = f"Creation datetime: " \
                                                               f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        writer.sheets['Summary'].cell(row=1, column=1).alignment = Alignment(horizontal='center')
        for sheet_name, dataframe in zip(devs_names, flatsDataframes):
            dataframe.to_excel(writer, sheet_name)
            auto_adjust_xlsx_column_width(dataframe, writer, sheet_name=sheet_name, margin=1)
    memory_file.seek(0)
    file = memory_file.getbuffer().tobytes()
    return file


def create_summary_to_excel(dataframes: list, developer_names: list):
    summary = []
    for dataframe, developer_name in zip(dataframes, developer_names):
        summary.append({"developer": developer_name})
        investments = dataframe.groupby(by="invest_name")
        for investment in investments:
            currentInvestment = investments.get_group(investment[0])
            flatsStatus = currentInvestment['status'].value_counts(dropna=False).to_dict()
            if flatsStatus.get('wolne') is None:
                flatsStatus['wolne'] = 0
            if flatsStatus.get(None) is None:
                flatsStatus[None] = 0
            if flatsStatus.get('zarezerwowane') is None:
                flatsStatus['zarezerwowane'] = 0
            if flatsStatus.get('sprzedane') is None:
                flatsStatus['sprzedane'] = 0
            flatsToPricePerSqm = currentInvestment[
                currentInvestment['area'].notnull() & currentInvestment['status'].ne("sprzedane")]
            investmentSummary = {
                "investment_name": investment[0],
                "amount_of_all_flats": len(currentInvestment),
                "average_min_price_per_m2": round((flatsToPricePerSqm['price'] / flatsToPricePerSqm['area']).min(), 2),
                "average_max_price_per_m2": round((flatsToPricePerSqm['price'] / flatsToPricePerSqm['area']).max(), 2),
                "amount_of_free_flats": flatsStatus['wolne'] + flatsStatus[None],
                "amount_of_reserved_flats": flatsStatus.get('zarezerwowane'),
                "amount_of_sold_flats": flatsStatus.get('sprzedane'),
                "percentage_to_sold/all_flats": round(
                    (len(currentInvestment) - flatsStatus['wolne'] - flatsStatus[None] -
                     flatsStatus['zarezerwowane']) / len(currentInvestment), 2)
            }
            summary.append(investmentSummary)
    summaryDf = pd.DataFrame(summary)
    return summaryDf
